from tkinter import *
from datetime import date
from tkinter import filedialog, messagebox
from tkinter.ttk import Combobox
from PIL import Image, ImageTk
import os
import openpyxl
from openpyxl import Workbook
import pathlib
import shutil

bg_main = "#f4f6f8"
frame_bg = "#ffffff"
frame_fg = "#344955"
btn_primary = "#4a90e2"
btn_secondary = "#f76c6c"

file = pathlib.Path("student_data.xlsx")
if not file.exists():
    wb = Workbook()
    sheet = wb.active
    headers = [
        "Registration No.",
        "Name",
        "Class",
        "Gender",
        "DOB",
        "Date Of Registration",
        "Religion",
        "Skill",
        "Father Name",
        "Mother Name",
        "Father's Occupation",
        "Mother's Occupation",
        "Photo",
    ]
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)
    wb.save("student_data.xlsx")

photos_dir = "photos"
if not os.path.exists(photos_dir):
    os.makedirs(photos_dir)

photo_path = None


# selection
def selection():
    return "Male" if radio.get() == 1 else "Female"


# Exit
def Exit():
    root.destroy()


# show_image
def show_image():
    global photo_path
    file_path = filedialog.askopenfilename(
        title="Select Image", filetypes=[("Image Files", "*.jpg *.png *.jpeg")]
    )
    if file_path:
        img_open = Image.open(file_path)
        img_open = img_open.resize((150, 150))
        img_tk = ImageTk.PhotoImage(img_open)
        img_label.config(image=img_tk)
        img_label.image = img_tk
        file_name = os.path.basename(file_path)
        save_path = os.path.join(photos_dir, file_name)
        shutil.copy(file_path, save_path)
        photo_path = save_path


# save_student
def save_student():
    global photo_path
    reg_no = Registration.get()
    name = Name.get()
    student_class = Class.get()
    gender = selection()
    dob = DOB.get()
    reg_date = Date.get()
    religion = Religion.get()
    skill = Skills.get()
    father = Father_Name.get()
    mother = Mother_Name.get()
    father_occ = Father_Occupation.get()
    mother_occ = Mother_Occupation.get()

    if not all(
        [
            reg_no,
            name,
            student_class,
            gender,
            dob,
            reg_date,
            religion,
            skill,
            father,
            mother,
            father_occ,
            mother_occ,
            photo_path,
        ]
    ):
        messagebox.showerror("Error", "Please fill all fields and upload a photo.")
        return

    wb = openpyxl.load_workbook("student_data.xlsx")
    sheet = wb.active
    sheet.append(
        [
            reg_no,
            name,
            student_class,
            gender,
            dob,
            reg_date,
            religion,
            skill,
            father,
            mother,
            father_occ,
            mother_occ,
            photo_path,
        ]
    )
    wb.save("student_data.xlsx")
    messagebox.showinfo("Success", "Student data saved successfully.")


# reset_form
def reset_form():
    global photo_path
    Registration.set("")
    Name.set("")
    DOB.set("")
    radio.set(0)
    Class.set("Select Class")
    Religion.set("")
    Skills.set("")
    Father_Name.set("")
    Father_Occupation.set("")
    Mother_Name.set("")
    Mother_Occupation.set("")
    img_label.config(image="", text="", bg="black")
    img_label.image = None
    photo_path = None
    Date.set(date.today().strftime("%d/%m/%Y"))


root = Tk()
root.title("Student Registration System")
root.geometry("1000x600+250+120")
root.config(bg=bg_main)

Label(
    root,
    text="Email: Mukhtarrahimi110@gmail.com",
    bg=btn_secondary,
    fg="white",
    font=("Arial", 10),
    anchor="e",
).pack(side=TOP, fill=X)
Label(
    root,
    text="STUDENT REGISTRATION",
    bg=frame_fg,
    fg="white",
    font=("Arial", 18, "bold"),
).pack(side=TOP, fill=X)

search = StringVar()
Entry(root, textvariable=search, width=18, bd=2, font=("Arial", 14)).place(x=650, y=60)
srch = Button(
    root, text="Search", bg=btn_primary, fg="white", font=("Arial", 12, "bold")
)
srch.place(x=850, y=57)

update_button = Button(
    root, text="Update", bg=btn_secondary, fg="white", font=("Arial", 12, "bold")
)
update_button.place(x=100, y=57)

Label(root, text="Registration No:", bg=bg_main, fg=frame_fg, font=("Arial", 11)).place(
    x=30, y=120
)
Label(root, text="Date:", bg=bg_main, fg=frame_fg, font=("Arial", 11)).place(
    x=400, y=120
)

Registration = StringVar()
Date = StringVar()
Entry(root, textvariable=Registration, width=15, font=("Arial", 10)).place(x=150, y=120)

today = date.today()
Date.set(today.strftime("%d/%m/%Y"))
Entry(root, textvariable=Date, width=15, font=("Arial", 10)).place(x=450, y=120)

obj = LabelFrame(
    root,
    text="Student Details",
    bg=frame_bg,
    fg=frame_fg,
    font=("Arial", 14, "bold"),
    width=940,
    height=200,
)
obj.place(x=30, y=160)

Label(obj, text="Name:", bg=frame_bg, fg=frame_fg, font=("Arial", 11)).place(x=20, y=30)
Label(obj, text="Date of Birth:", bg=frame_bg, fg=frame_fg, font=("Arial", 11)).place(
    x=20, y=70
)
Label(obj, text="Gender:", bg=frame_bg, fg=frame_fg, font=("Arial", 11)).place(
    x=20, y=110
)

Label(obj, text="Class:", bg=frame_bg, fg=frame_fg, font=("Arial", 11)).place(
    x=400, y=30
)
Label(obj, text="Religion:", bg=frame_bg, fg=frame_fg, font=("Arial", 11)).place(
    x=400, y=70
)
Label(obj, text="Skills:", bg=frame_bg, fg=frame_fg, font=("Arial", 11)).place(
    x=400, y=110
)

Name = StringVar()
Entry(obj, textvariable=Name, width=18, font=("Arial", 10)).place(x=120, y=30)

DOB = StringVar()
Entry(obj, textvariable=DOB, width=18, font=("Arial", 10)).place(x=120, y=70)

radio = IntVar()
Radiobutton(
    obj,
    text="Male",
    variable=radio,
    value=1,
    bg=frame_bg,
    fg=frame_fg,
    command=selection,
).place(x=120, y=110)
Radiobutton(
    obj,
    text="Female",
    variable=radio,
    value=2,
    bg=frame_bg,
    fg=frame_fg,
    command=selection,
).place(x=180, y=110)

Class = Combobox(obj, values=[str(i) for i in range(1, 13)], width=15)
Class.place(x=500, y=30)
Class.set("Select Class")

Religion = StringVar()
Entry(obj, textvariable=Religion, width=18, font=("Arial", 10)).place(x=500, y=70)

Skills = StringVar()
Entry(obj, textvariable=Skills, width=18, font=("Arial", 10)).place(x=500, y=110)

obj2 = LabelFrame(
    root,
    text="Parents Details",
    bg=frame_bg,
    fg=frame_fg,
    font=("Arial", 14, "bold"),
    width=940,
    height=150,
)
obj2.place(x=30, y=370)

Label(obj2, text="Father's Name:", bg=frame_bg, fg=frame_fg, font=("Arial", 11)).place(
    x=20, y=30
)
Label(obj2, text="Occupation:", bg=frame_bg, fg=frame_fg, font=("Arial", 11)).place(
    x=20, y=70
)
Label(obj2, text="Mother's Name:", bg=frame_bg, fg=frame_fg, font=("Arial", 11)).place(
    x=400, y=30
)
Label(obj2, text="Occupation:", bg=frame_bg, fg=frame_fg, font=("Arial", 11)).place(
    x=400, y=70
)

Father_Name = StringVar()
Entry(obj2, textvariable=Father_Name, width=18, font=("Arial", 10)).place(x=140, y=30)

Father_Occupation = StringVar()
Entry(obj2, textvariable=Father_Occupation, width=18, font=("Arial", 10)).place(
    x=140, y=70
)

Mother_Name = StringVar()
Entry(obj2, textvariable=Mother_Name, width=18, font=("Arial", 10)).place(x=520, y=30)

Mother_Occupation = StringVar()
Entry(obj2, textvariable=Mother_Occupation, width=18, font=("Arial", 10)).place(
    x=520, y=70
)

f = Frame(root, bd=2, bg="black", width=150, height=150)
f.place(x=800, y=160)

img_label = Label(f, bg="black")
img_label.pack()

Button(
    root,
    text="Upload Photo",
    bg=btn_primary,
    fg="white",
    font=("Arial", 11, "bold"),
    width=15,
    command=show_image,
).place(x=800, y=320)

Button(
    root,
    text="Save",
    bg=btn_primary,
    fg="white",
    font=("Arial", 11, "bold"),
    width=15,
    command=save_student,
).place(x=800, y=370)
Button(
    root,
    text="Reset",
    bg="#ffa534",
    fg="white",
    font=("Arial", 11, "bold"),
    width=15,
    command=reset_form,
).place(x=800, y=420)
Button(
    root,
    text="Exit",
    bg=btn_secondary,
    fg="white",
    font=("Arial", 11, "bold"),
    width=15,
    command=Exit,
).place(x=800, y=470)

root.mainloop()
